import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import config
from config import get_cloud_name

class Rule:
    def __init__(self, rule_id, rule_type, action, enabled, logging, orig_ip, orig_port, trans_ip, trans_port, protocol, description):
        self.rule_id = rule_id
        self.rule_type = rule_type
        self.action = action
        self.enabled = enabled
        self.logging = logging
        self.orig_ip = orig_ip
        self.orig_port = orig_port
        self.trans_ip = trans_ip
        self.trans_port = trans_port
        self.protocol = protocol
        self.description = description

def parse_rules(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    rules = []

    for rule in root.findall('.//rule'):
        rule_id = rule.find('ruleId').text
        rule_type = rule.find('ruleType').text
        action = rule.find('action').text
        enabled = rule.find('enabled').text
        logging = rule.find('loggingEnabled').text
        orig_ip = rule.find('originalAddress').text
        orig_port = rule.find('originalPort').text
        trans_ip = rule.find('translatedAddress').text
        trans_port = rule.find('translatedPort').text
        protocol = rule.find('protocol').text
        desc = rule.find('description')
        description = desc.text if desc is not None else "Отсутствует"

        rules.append(Rule(rule_id, rule_type, action, enabled, logging, orig_ip, orig_port, trans_ip, trans_port, protocol, description))

    return rules

def rules_to_dataframe(rules):
    data = {
        'ID': [rule.rule_id for rule in rules],
        'Type': [rule.rule_type for rule in rules],
        'Action': [rule.action for rule in rules],
        'Enabled': [rule.enabled for rule in rules],
        'Logging': [rule.logging for rule in rules],
        'OrigIP': [rule.orig_ip for rule in rules],
        'OrigPort': [rule.orig_port for rule in rules],
        'TransIP': [rule.trans_ip for rule in rules],
        'TransPort': [rule.trans_port for rule in rules],
        'Protocol': [rule.protocol for rule in rules],
        'Description': [rule.description for rule in rules]
    }
    return pd.DataFrame(data)

def clean_data(df):
    df.replace(regex={r'<[^>]+>': ' ', r'\s(false|true)\s': ''}, inplace=True)
    df['Description'] = df['Description'].str.lstrip()
    return df

def save_to_excel(df, file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # Очистка листа
    sheet.delete_rows(1, sheet.max_row)

    # Запись заголовков
    for col_num, column_title in enumerate(df.columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Запись данных
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Добавление фильтрации
    full_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    sheet.auto_filter.ref = full_range

    wb.save(file_path)

def main():
    file_path = config.input_file
    rules = parse_rules(file_path)
    df = rules_to_dataframe(rules)
    df = clean_data(df)

    cloud_name = get_cloud_name()
    if cloud_name:
        print(f"Выбрано облако: {cloud_name}")
    else:
        print("Неверный выбор облака.")
        return

    save_to_excel(df, config.output_file, cloud_name)
    print('Файл готов')

if __name__ == "__main__":
    main()